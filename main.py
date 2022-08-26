import openpyxl
wb = openpyxl.load_workbook('file.xlsx')
print(type(wb))
print(wb.sheetnames)
sheet = wb['S1']
print(sheet)
print(type(sheet))
print(sheet.title)
print(wb.active)
print("-------")
print(sheet['A1'])
print(sheet['A1'].value)
cell = sheet['B2']
print(cell.value)
print("Row " + str(cell.row) + ", column " + str(cell.column) + " has value " + str(cell.value) + ".")
print("Cell " + str(cell.coordinate) + ' has value ' + str(cell.value) + '.')
print("-------")
print(sheet.cell(row=1, column=2).value)
for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)
print("top row "+ str(sheet.max_row)+ " top column "+str(sheet.max_column))
print(openpyxl.utils.get_column_letter(885))
print(openpyxl.utils.column_index_from_string('AA'))
print(tuple(sheet['A1':'C3']))
for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- enter ---')

for rowOfCellObjects in sheet['A1':'C3']:
    row=""
    for cellObj in rowOfCellObjects:
        row=row+str((cellObj.coordinate, cellObj.value))
    print(row)
print()
for rowOfCellObjects in sheet['A1':'C3']:
    row=""
    for cellObj in rowOfCellObjects:
        row=row+str(cellObj.value)+" "
    print(row)
print("-------")
for cell in sheet['a']:
   print (str(cell.value))
sheet.title = 'cheese ;)'
#wb.save('example_copy.xlsx')
print("test")
print(wb.sheetnames)
print(wb.create_sheet())
print(wb.sheetnames)
print(wb.create_sheet(index=0, title='First one'))
print(wb.sheetnames)
del wb["Sheet"]
print(wb.sheetnames)
sheet2 = wb['First one']
sheet2['A1'] = 'Hello!'#
#sheet2['A1'].style.font.size = 22
#sheet2['A1'].font = Font(size=20)
#sheet2['A1'].style = (size=24, italic=True)
wb.save('example_copy.xlsx')

